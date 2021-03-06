#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
# @Time    : 19/11/11 13:27
# @Author  : Lei Zhen
# @Contract: leizhen8080@gmail.com
# @Site    : http://www.leizhen.com
# @File    : extract_promotion_detail.py
# @Software: PyCharm
# code is far away from bugs with the god animal protecting
    I love animals. They taste delicious.
             ┏┓   ┏┓
            ┏┛┻━━━┛┻┓
            ┃  ☃    ┃
            ┃ ┳┛  ┗┳ ┃
            ┃      ┻  ┃
            ┗━┓     ┏━┛
              ┃     ┗━━━┓
              ┃ 神兽保佑 ┣┓
              ┃ 永无BUG┏┛
              ┗┓┓┏━┳┓┏┛
               ┃┫┫ ┃┫┫
               ┗┻┛ ┗┻┛
"""
import json
import os
import re
import time
import traceback
from datetime import datetime

from openpyxl import Workbook

from douyin.conver_encoding import convert_file, clean_dir
from tools.logger import logger_wei as logger
from tools.parser_config import get_db_tool

log = logger


def create_excel(file_path):
    wb = Workbook(write_only=True)
    wb.create_sheet('素人账号信息')
    wb.create_sheet('商品橱窗信息')
    wb.create_sheet('账号运营信息')
    wb.save(file_path)


def format_promotion(json_data):
    result = json.loads(json_data)
    promotion_count = result.get('count')  # 商品总数
    promotions = result.get('promotions')
    logger.debug('商品信息')
    logger.debug(promotions)
    for promotion in promotions:
        sales = promotion.get('sales')
        title = promotion['title']
        visitor = promotion.get('visitor')
        if visitor:
            visitor_count = visitor.get('count')
        else:
            visitor_count = None
        price = str(promotion.get('price', 0) / 100.0)
        last_aweme_id = promotion.get('last_aweme_id')  # 对应视频id
        promotion_id = promotion['promotion_id']
        goods_source = promotion.get('goods_source', '')
        coupon_url, coupon_real_price = '', ''
        coupon = promotion.get('apply_coupon_button')
        if coupon:
            coupon_url = coupon.get('h5_url')
            coupon_real_price = re.findall(r'\d+\.*\d*', coupon.get('text')[0])[0]
        elastic_title = promotion.get('elastic_title')
        elastic_type = promotion.get('elastic_type')
        detail_url = promotion.get('detail_url')
        # line = f'名称：{title}，\n\t销量：{sales}，访客：{visitor_count}，价格：{price}'
        # print(line)
        promotion_dict = {'商品id': promotion_id, '名称': title, '销量': sales, '访客': visitor_count, '价格': price,
                          '商品来源': goods_source, '优惠后价格': coupon_real_price, '弹性标题': elastic_title,
                          '弹性类型': elastic_type, '关联视频id': last_aweme_id, '优惠券链接': coupon_url,
                          '商品详情链接': detail_url, '商品总数': promotion_count}
        logger.info('商品信息-提取内容')
        logger.info(promotion_dict)
        yield promotion_dict


def format_user_info(json_data):
    result = json.loads(json_data)
    log.debug('用户信息')
    logger.debug(result)
    user = result['user']
    uid = user.get('uid')
    total_favorited = user.get('total_favorited')  # 获赞量
    following_count = user.get('following_count')  # 关注量
    follower_count = user.get('follower_count')  # 粉丝量
    aweme_count = user.get('aweme_count')  # 作品数
    dongtai_count = user.get('dongtai_count')  # 动态数
    favoriting_count = user.get('favoriting_count')  # 喜欢量
    bind_phone = user.get('bind_phone')  # 绑定手机
    bind_phone = bind_phone if bind_phone != '#' else ''
    signature = user.get('signature')  # 签名，可能包含手机号和微信号
    nickname = user.get('nickname')  # 昵称
    birthday = user.get('birthday')  # 出生日期
    age = None  # 年龄
    if birthday:
        # 计算年龄
        now = datetime.now().year
        birthday_year = int(birthday[:4])
        age = now - birthday_year
    city = user.get('city')  # 城市
    school_name = user.get('school_name')
    college_name = user.get('college_name')  # 大学
    if not school_name:
        school_name = college_name
    gender = user.get('gender')  # 性别，1：男，2：女
    if gender:
        gender = '男' if int(gender) == 1 else '女'
    unique_id = user.get('unique_id', '')  # 抖音号
    short_id = user.get('short_id', '')  # 抖音号2
    if not unique_id:
        unique_id = short_id
    with_fusion_shop_entry = user.get('with_fusion_shop_entry')  # 是否有商品橱窗
    if with_fusion_shop_entry is not None:
        with_fusion_shop_entry = int(with_fusion_shop_entry)
    custom_verify = user.get('custom_verify')  # 自定义认证
    enterprise_verify_reason = user.get('enterprise_verify_reason')  # 官方认证

    # user_url = user['share_info']['share_url']  # 用户详情链接
    # user_qr_code_url = user['share_info']['share_qrcode_url']['url_list']
    # if len(user_qr_code_url) > 0:
    #     user_qr_code_url = user_qr_code_url[0]

    user_dict = {'素人id': uid, '': unique_id, '获赞量': total_favorited, '关注量': following_count, '作品数': aweme_count,
                 '动态数': dongtai_count, '粉丝量': follower_count, '喜欢量': favoriting_count, '绑定手机': bind_phone,
                 '签名': signature, '昵称': nickname, '年龄': age, '性别': gender, '城市': city, '大学': school_name,
                 '是否有商品橱窗': with_fusion_shop_entry, '黄标': custom_verify, '蓝标': enterprise_verify_reason,
                 '用户详情链接': None, '用户详情二维码链接': None}
    logger.info('用户信息-提取内容')
    logger.info(user_dict)
    yield user_dict


def form_aweme(json_data):
    data = json.loads(json_data)
    aweme_list = data['aweme_list']
    logger.debug('作品信息')
    logger.debug(aweme_list)
    for aweme in aweme_list:
        # 作品详情
        statistic = aweme['statistics']
        digg_count = statistic['digg_count']  # 点赞量
        comment_count = statistic['comment_count']  # 评论量
        share_count = statistic['share_count']  # 分享量
        download_count = statistic['download_count']  # 下载量
        forward_count = statistic['forward_count']  # 转发量
        aweme_id = aweme['aweme_id']  # 作品id
        author_user_id = str(aweme['author_user_id'])  # 用户id
        simple_promotions = aweme.get('simple_promotions')
        # 获取商品id和另一个product_id
        promotion_id, product_id = '', ''
        if simple_promotions is not None:
            promotions_dict = json.loads(simple_promotions)
            if promotions_dict is not None and len(promotions_dict) > 0:
                promotions_dict = promotions_dict[0]
                promotion_id = promotions_dict.get('promotion_id', '')
                product_id = promotions_dict.get('product_id', '')
        # 分类
        text_extra = aweme['text_extra']
        tag = []
        for text in text_extra:
            if 'hashtag_name' in text:
                tag.append(text['hashtag_name'])
        tag = ','.join(tag)  # 标签
        desc = aweme['desc']  # 视频标题
        create_time = datetime.fromtimestamp(int(aweme['create_time']))  # 创建时间
        aweme_dict = {'视频id': aweme_id, '素人id': author_user_id, '点赞量': digg_count, '评论量': comment_count,
                      '分享量': share_count, '下载量': download_count, '转发量': forward_count, '商品id': promotion_id,
                      '产品id': product_id, '标签': tag, '视频标题': desc, '发布时间': create_time}
        logger.info('作品信息-提取内容')
        logger.info(aweme_dict)
        yield aweme_dict


source_base_dir = 'D:\\douyin'
base_dir = 'D:\\douyin2'
all_files = os.listdir('D:\\douyin2')
if not all_files:
    all_files = os.listdir(base_dir)

# 获取配置文件
env_file = os.path.abspath(os.path.join(os.path.relpath(__file__), os.path.pardir, '.env'))


def handle_file(files):
    db_pool = get_db_tool()
    conn = db_pool.connection()
    cursor = conn.cursor()

    # 查询 作品数据
    # aweme_query_sql = 'SELECT aweme_id FROM douyin_aweme WHERE DATEDIFF(update_time, NOW()) = 0'
    # cursor.execute(query=aweme_query_sql)
    # today_aweme_result = cursor.fetchall()
    # today_aweme_result = [item[0] for item in today_aweme_result]

    # 查询 商品数据
    # promotion_query_sql = "SELECT promotion_id FROM douyin_promotion WHERE DATEDIFF(update_time, NOW()) = 0"
    # cursor.execute(promotion_query_sql)
    # today_promotion_result = cursor.fetchall()
    # today_promotion_result = [item[0] for item in today_promotion_result]

    for file in files:
        file_dir = os.path.join(base_dir, file)
        file_name = os.path.splitext(file)[0]
        if 'promotion' in file_name:
            user_id = re.match(r'(\d+)', file_name).groups()[0]
            with open(file_dir, 'rt', encoding='utf-8') as f:
                for line in f:
                    line = line.strip().strip('\n')
                    if not line:
                        logger.info(f'文件当前行 {file} 为空，请检查')
                        continue
                    promotion_infos = list(format_promotion(line))
                    # new_promotion_infos = list(
                    #     filter(lambda x: x['商品id'] not in today_promotion_result, promotion_infos))
                    promotion_insert_args = [
                        list(item.values()) for item in promotion_infos]  # todo：需要删除
                    for x in promotion_insert_args:
                        x.insert(1, user_id)
                    # promotion_insert_args = [list(item.values()).insert(1, user_id) for item in new_promotion_infos]
                    logger.info(f'新增的商品:{[item[0] for item in promotion_insert_args]}')
                    # 批量插入数据
                    promotion_insert_sql = """REPLACE INTO douyin_promotion2
        (promotion_id, suren_id, title, sales, visitor_count, price, goods_source, 
        coupon_real_price, elastic_title, elastic_type,
         aweme_id, coupon_link, detail_url, count, update_time)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, NOW())"""
                    try:
                        cursor.executemany(promotion_insert_sql, promotion_insert_args)
                        conn.commit()
                    except Exception as e:
                        logger.error(traceback.format_exc())
                        conn.rollback()

                # 写入 excel
                # df = pd.DataFrame(list(format_promotion(line)))
                # user_id = re.match(r'(\d+)', file_name).groups()[0]
                # writer = pd.ExcelWriter(f'{user_id}_promotion.xlsx')
                # df.to_excel(writer, sheet_name='商品橱窗')
                # writer.close()
        elif 'user' in file_name:
            with open(file_dir, 'rt', encoding='utf-8') as f:
                for line in f:
                    line = line.strip().strip('\n')
                    if not line:
                        logger.info(f'文件当前行 {file_name} 为空')
                        continue
                    if '参数不合法' in line:
                        logger.error(f'文件当前行 {file_name} 参数不合法，请检查: {line}')
                        continue
                    suren_infos = list(format_user_info(line))
                    if suren_infos is None or len(suren_infos) < 1:
                        logger.error('从文件获取用户失败')
                        continue

                    suren_info = suren_infos[0]

                    suren_id = suren_info.get('素人id')
                    # 如果该素人数据已经存在，则不存入数据库
                    # 后期可以改为新增 create_time
                    suren_query = 'SELECT * FROM douyin_user WHERE suren_id = %s '
                    cursor.execute(query=suren_query, args=suren_id)
                    result = cursor.fetchone()
                    if result is not None:
                        logger.error('用户不是新用户')
                        continue
                    insert_args = [list(suren_info.values()) for suren_info in suren_infos]
                    # 如果没有当天的数据则存入/更新数据库
                    suren_insert = """REPLACE INTO douyin_user
                (suren_id, douyin_id,total_favourited, following_count, aweme_count, dongtai_count, follower_count,
                 favouriting_count, bind_phone, signature, nickname, age, gender, city, college, has_promotion, 
                 custom_verify, enterprise_verify_reason, url, qrcode_url, update_time)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, NOW())"""
                    try:
                        cursor.executemany(query=suren_insert, args=insert_args)
                        conn.commit()
                    except Exception:
                        logger.error(traceback.format_exc())
                        conn.rollback()
                        # df = pd.DataFrame(list(format_user_info(line)))
                        # user_id = df['素人id'].values[0]
                        # # if not os.path.exists(f'{user_id}.xls'):
                        # writer = pd.ExcelWriter(f'{user_id}_user.xlsx')
                        # df.to_excel(writer, sheet_name='素人信息')
                        # writer.close()
        elif 'zuopin' in file_name:
            with open(file_dir, 'rt', encoding='utf-8') as f:
                for line in f:
                    line = line.strip().strip('\n')
                    if not line:
                        logger.info(f'文件当前行 {file_name} 为空，请检查')
                        continue
                    aweme_infos = list(form_aweme(line))
                    # 求差集，获取今天没有存储过的作品id
                    # new_aweme_infos = list(filter(lambda info: info['视频id'] not in today_aweme_result, aweme_infos))
                    new_aweme_args = [list(item.values()) for item in aweme_infos]  # todo 需要删除
                    # new_aweme_args = [list(item.values()) for item in new_aweme_infos]
                    logger.info(f'本次录入的新作品，id： {[item[0] for item in new_aweme_args]}')
                    aweme_insert_sql = """REPLACE INTO douyin_aweme2
        (aweme_id, suren_id, digg_count, comment_count, share_count, download_count,
         forward_count, promotion_id, product_id, tag, description, create_time, update_time)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, NOW())"""
                    try:
                        cursor.executemany(aweme_insert_sql, new_aweme_args)
                        conn.commit()
                    except:
                        logger.error(traceback.format_exc())
                        conn.rollback()
            # df = pd.DataFrame(list(form_aweme(line)))
            # user_id = df['素人id'].values[0]
            # df.to_excel(f'{user_id}_zuopin.xlsx', sheet_name='作品信息')

    conn.close()


# # 转换 utf-16 为 utf-8
def extract():
    success = convert_file()
    handle_file(os.listdir(base_dir))
    clean_dir(base_dir)
    clean_dir(source_base_dir)


if __name__ == '__main__':
    extract()

# # todo: 不知道能不能做去重，先清空数据再说
# def clean():
#     all_file = os.listdir(base_dir)
#     for file in all_file:
#         os.remove(file)
